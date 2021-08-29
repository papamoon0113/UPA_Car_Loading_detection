import openpyxl
import count_mgmt, history_mgmt
from openpyxl.utils import get_column_letter
import pandas as pd

def find(st, value):
    for j in range(1, 10):
            for i in range(1, st.max_row+1):
                col = get_column_letter(j)
                if st[col + str(i)].value == value:
                    # print(col + str(i))
                    return [j, str(i)]
    print(f'{value} is not in {st}!!')

def update_xlsx(xlsx, m_from, m_to, moving_time):
    wb = openpyxl.load_workbook(xlsx)
    st = wb['Sheet1']

    translate = {'STF':'SantaFeTM', 'AVT':'AvanteMD', 'i30':'i30', 'NKON':'TheNewKona', 'NSR':'NewSorentoR', 'TNC':'TheNewCanival', 'LFS':'LFSonata', 'NFS':'NFSonata', 'YFS':'YFSonata', 'GS':'GrandStarex', 'GHG':'GranduerHG'}

    # 엑셀파일에서 total이라는 글자 찾기
    location = find(st, 'TOTAL')
    location_model = find(st, "DEST'N ; POSTS / STATES")

    # 그 옆으로 계속 이동하면서 숫자면 위의 칸 중 글자인 것을 찾기
    _count = count_mgmt.Count()
    _history = history_mgmt.History()

    for i in range(location[0]+1, st.max_column):
        col = get_column_letter(i)
        if type(st[col +location[1]].value) == int:
            count = st[col + location[1]].value
            if st[col + location_model[1]].value == 'UNIT':
                break
            else:
                value = translate[st[col + location_model[1]].value]
                # moving_history에 반영
                _history.create(value, m_from, m_to, count, moving_time)

                # model_count에 반영
                if m_from == 'YARD':
                    _count.minus_count(value, 'YARD', count)
                if m_from == '6wharf':
                    _count.minus_count(value, '6wharf', count)
                if m_to == 'YARD':
                    _count.plus_count(value, 'YARD', count)
                if m_to == '6wharf':
                    _count.plus_count(value, '6wharf', count)
    # 그 글자를 위의 translate로 변환해서 create 하기



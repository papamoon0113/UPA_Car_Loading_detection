import openpyxl
from openpyxl.styles.borders import Border
import pandas as pd
from mysql_util import Mysql
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment

def make_summary():
    wb = openpyxl.load_workbook('template.xlsx')
    st = wb['Sheet1']

    mysql = Mysql()
    mysql.set_table('car_carrier_tally')
    car_df = mysql.get_data()
    mysql.set_table('car_info')
    car_info_df = mysql.get_data()

    def get_weight(car_info_df, car):
        return float(car_info_df[car_info_df['MODEL']==car]['WEIGHT'].iat[0])

    def get_cubic(car_info_df, car):
        return float(car_info_df[car_info_df['MODEL']==car]['CUBIC'].iat[0])

    columns = ['post', 'state', 'VSR', 'AVT', 'AVTH', 'I30', 'IOH', 'IOE5', 'G70', 'G80', 'G90', 'VNU', 'KON', 'KOH', 'NKON', 'KOE', 'TUC', 'TUCH', 'STF', 'STFH', 'PSD', 'GV70', 'GV80', 'PTR', 'G80E', 'unit', 'weight', 'cubic']

    col_letter = {}
    i=2
    for col in columns:
        col_letter[col] = get_column_letter(i)
        i+=1


    thin = Side(border_style="thin", color="000000")
    row = 7
    count = 0
    weight = 0
    cubic = 0
    pol_ = True
    for pol in car_df.POL.unique():
        df = car_df[car_df.POL == pol]
        if pol_:
            st[col_letter['post']+str(row)] = pol
            pol_ = False
        
        for pod in df.POD.unique():
            st[col_letter['state']+str(row)] = pod
            df_ = df[df.POD == pod]
            df_.reset_index(inplace=True)
            del df_['index']
            for i in range(len(df_)):
                model = df_.loc[i]['MODEL']
                col = col_letter[model]
                if not st[col+str(row)].value:
                    st[col+str(row)] = 1
                    count +=1
                    # weight += get_weight(car_info_df, model)
                    weight += get_cubic(car_info_df, model)
                    # cubic += get_cubic(car_info_df, model)
                    cubic += get_weight(car_info_df, model)
                else:
                    st[col+str(row)] = st[col+str(row)].value + 1
                    count +=1
                    # weight += get_weight(car_info_df, model)
                    weight += get_cubic(car_info_df, model)
                    # cubic += get_cubic(car_info_df, model)
                    cubic += get_weight(car_info_df, model)

            st[col_letter['unit']+str(row)] = count
            st[col_letter['weight']+str(row)] = weight
            st[col_letter['cubic']+str(row)] = cubic
            count = 0
            weight = 0
            cubic = 0
            row +=1
        pol_ = True

    #밑의 토탈 값 출력
    tot_row = row
    value_range = range(7, row)
    st['B'+str(row)] = 'TOTAL'
    st['B'+str(row)].alignment = Alignment(horizontal='center')
    st.merge_cells(f"{'B'+str(row)}:{'C'+str(row)}")

    value_col = list(col_letter.values())

    for col in value_col[2:]:
        tot = 0
        for i in value_range:
            if not st[col + str(i)].value:
                continue
            else:
                tot+=st[col + str(i)].value
        if tot != 0:
            st[col+str(row)] = tot
        tot = 0

    bold = Font(bold=True)
    for col in value_col:
        st[col + str(row)].font = bold

    row +=1
    for col in value_col:
        for i in range(7, row):
            st[col + str(i)].border = Border(top=thin, bottom=thin, left=thin, right=thin)

    result_row = row
    for pol in car_df.POL.unique():
        st['B'+str(row)] = pol
        st.merge_cells(f"{'B'+str(row)}:{'C'+str(row)}")
        df = car_df[car_df.POL == pol]
        df['count'] = 1
        st[value_col[2]+str(row)] = df['count'].sum()
        st.merge_cells(f"{value_col[2]+str(row)}:{value_col[3]+str(row)}")
        st[value_col[4]+str(row)] = 'UNITS'
        
        df.reset_index(inplace=True)
        del df['index']

        for i in range(len(df)):
            model = df.loc[i]['MODEL']
            # weight += get_weight(car_info_df, model)
            weight += get_cubic(car_info_df, model)
            # cubic += get_cubic(car_info_df, model)
            cubic += get_weight(car_info_df, model)

        st[value_col[6]+str(row)] = weight
        st.merge_cells(f"{value_col[6]+str(row)}:{value_col[7]+str(row)}")
        st[value_col[8]+str(row)] = 'K/TON'

        st[value_col[9]+str(row)] = cubic
        st.merge_cells(f"{value_col[9]+str(row)}:{value_col[10]+str(row)}")
        st[value_col[11]+str(row)] = 'CBM'
        
        row +=1
        weight = 0
        cubic = 0

    st['B'+str(row)] = 'G / TOTAL'
    st['B'+str(row)].alignment = Alignment(horizontal='center')
    st.merge_cells(f"{'B'+str(row)}:{'C'+str(row)}")
    st[value_col[2]+str(row)] = st[value_col[-3]+str(tot_row)].value
    st.merge_cells(f"{value_col[2]+str(row)}:{value_col[3]+str(row)}")
    st[value_col[4]+str(row)] = 'UNITS'
    st[value_col[6]+str(row)] = st[value_col[-2]+str(tot_row)].value
    st.merge_cells(f"{value_col[6]+str(row)}:{value_col[7]+str(row)}")
    st[value_col[8]+str(row)] = 'K/TON'
    st[value_col[9]+str(row)] = st[value_col[-1]+str(tot_row)].value
    st.merge_cells(f"{value_col[9]+str(row)}:{value_col[10]+str(row)}")
    st[value_col[11]+str(row)] = 'CBM'

    for col in value_col[:12]:
        for i in range(tot_row+1, row):
            st[col + str(i)].border = Border(top=thin, bottom=thin)

    bd = Side(style='thick', color="000000")
    for col in value_col[:12]:
        st[col+str(row)].border = Border(top=bd, bottom=bd)
        st[col+str(row)].font = bold


    wb.save('result.xlsx')

make_summary()



